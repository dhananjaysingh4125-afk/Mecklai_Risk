from scipy.stats import norm
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots


import pandas as pd
import requests
from datetime import date

def get_live_rates():
    # Free Tier: exchangeratesapi.io or frankfurter.app
    url   = "https://api.frankfurter.app/latest?from=USD&to=INR,EUR,GBP,JPY"
    resp  = requests.get(url)
    rates = resp.json()["rates"]
    return rates

def load_positions(filepath):
    pos = pd.read_csv(filepath, parse_dates=["trade_date","value_date"])
    return pos[pos["value_date"] >= pd.Timestamp(date.today())]

rates     = get_live_rates()
positions = load_positions(r"C:\Users\91852\Desktop\Summers\SIP\Data.csv")
positions["mtm_inr"] = positions.apply(
    lambda r: r["notional_usd"] * rates.get(r["base_ccy"],1), axis=1)


limits = {"USD/INR": 5_000_000, "EUR/INR": 3_000_000, "GBP/INR": 2_000_000}

def rag_status(nop, limit):
    util = abs(nop) / limit
    if   util > 0.9: return "🔴 RED",   util
    elif util > 0.7: return "🟡 AMBER", util
    else:             return "🟢 GREEN", util

nop_by_ccy = positions.groupby("pair")["notional_usd"].sum()

for ccy, nop in nop_by_ccy.items():
    status, util = rag_status(nop, limits.get(ccy, 1e9))
    print(f"{ccy}: NOP={nop:,.0f} | Limit Util={util:.0%} | {status}")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Daily FX Risk Report"

# Header styling
header_fill = PatternFill("solid", fgColor="1a3a5c")
header_font = Font(color="FFFFFF", bold=True, size=11)

headers = ["Pair","NOP (USD)","Limit","Utilization %","Status","MTM P&L (INR)"]
for col, h in enumerate(headers, 1):
    cell          = ws.cell(row=1, column=col, value=h)
    cell.fill      = header_fill
    cell.font      = header_font
    cell.alignment = Alignment(horizontal="center")

wb.save(f"FX_Risk_Report_{date.today()}.xlsx")


import schedule
import time
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import date

def compute_risk_metrics(pos, rates):
    # MTM in INR
    pos["mtm_inr"] = pos.apply(
        lambda r: r["notional_usd"] * rates.get(r["base_ccy"], 1), axis=1)

    # Net Open Position by currency
    pos["nop"] = pos["notional_usd"]
    nop_by_ccy = pos.groupby("base_ccy")["nop"].sum().reset_index()

    # Limits
    limits = {"USD": 5_000_000, "EUR": 3_000_000, "GBP": 2_000_000, "JPY": 1_000_000}

    def rag(nop, ccy):
        limit = limits.get(ccy, 1e9)
        util  = abs(nop) / limit
        if   util > 0.90: return "RED",   round(util*100, 1)
        elif util > 0.70: return "AMBER", round(util*100, 1)
        else:             return "GREEN", round(util*100, 1)

    nop_by_ccy["status"], nop_by_ccy["util_pct"] = zip(
        *nop_by_ccy.apply(lambda r: rag(r["nop"], r["base_ccy"]), axis=1))

    # Daily PnL — MTM vs previous close (approximated as 0.5% move)
    pos["daily_pnl"] = pos["mtm_inr"] * 0.005

    return pos, nop_by_ccy


def generate_excel_report(pos, nop_by_ccy):
    wb = openpyxl.Workbook()

    # ── Sheet 1: NOP Summary ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "NOP Summary"

    header_fill = PatternFill("solid", fgColor="1a3a5c")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    red_fill    = PatternFill("solid", fgColor="FF4444")
    amber_fill  = PatternFill("solid", fgColor="FFA500")
    green_fill  = PatternFill("solid", fgColor="00AA44")

    headers = ["Currency", "NOP (USD)", "Limit (USD)", "Utilization %", "Status"]
    for col, h in enumerate(headers, 1):
        cell           = ws1.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    limits = {"USD": 5_000_000, "EUR": 3_000_000, "GBP": 2_000_000, "JPY": 1_000_000}

    for row_idx, row in enumerate(nop_by_ccy.itertuples(), 2):
        ws1.cell(row=row_idx, column=1, value=row.base_ccy)
        ws1.cell(row=row_idx, column=2, value=round(row.nop, 0))
        ws1.cell(row=row_idx, column=3, value=limits.get(row.base_ccy, "N/A"))
        ws1.cell(row=row_idx, column=4, value=f"{row.util_pct}%")

        status_cell      = ws1.cell(row=row_idx, column=5, value=row.status)
        status_cell.fill = {"RED": red_fill, "AMBER": amber_fill,
                            "GREEN": green_fill}[row.status]
        status_cell.font = Font(bold=True, color="FFFFFF")

    # ── Sheet 2: Position Detail ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Position Detail")
    detail_headers = ["Trade ID", "Client", "Pair", "Notional USD",
                      "Value Date", "MTM INR", "Daily PnL"]
    for col, h in enumerate(detail_headers, 1):
        cell      = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(pos.itertuples(), 2):
        ws2.cell(row=row_idx, column=1, value=row.trade_id)
        ws2.cell(row=row_idx, column=2, value=row.client)
        ws2.cell(row=row_idx, column=3, value=row.pair)
        ws2.cell(row=row_idx, column=4, value=round(row.notional_usd, 0))
        ws2.cell(row=row_idx, column=5, value=str(row.value_date.date()))
        ws2.cell(row=row_idx, column=6, value=round(row.mtm_inr, 0))
        ws2.cell(row=row_idx, column=7, value=round(row.daily_pnl, 0))

    # Auto-width
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[col[0].column_letter].width = max_len

    filename = f"FX_Risk_Report_{date.today()}.xlsx"
    wb.save(filename)
    print(f"✅ Report saved: {filename}")
    return filename


def run_daily_report():
    rates            = get_live_rates()
    pos              = load_positions(r"C:\Users\91852\Desktop\Summers\SIP\Data.csv")
    pos, nop_by_ccy  = compute_risk_metrics(pos, rates)   # unpack both outputs
    generate_excel_report(pos, nop_by_ccy)

if __name__ == "__main__":
    rates           = get_live_rates()
    pos             = load_positions(r"C:\Users\91852\Desktop\Summers\SIP\Data.csv")
    pos, nop_by_ccy = compute_risk_metrics(pos, rates)
    filename        = generate_excel_report(pos, nop_by_ccy)
    print(f"✅ Saved: {filename}")

import os
print(os.getcwd())          # shows current save location

# To save to a specific folder instead, change this line in generate_excel_report():
filename = f"C:/Users/91852/Desktop/Summers/SIP/FX_Risk_Report_{date.today()}.xlsx"
